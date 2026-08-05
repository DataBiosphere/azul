"""
Helpers for preparing a new anvilXY catalog release.

Workflow:

    1. Add rows for any new snapshots to the "AnVIL Prod" sheet of the "TDR
       Datasets and snapshots" spreadsheet, plus an empty ``anvilXY`` header
       cell if not already present.

    2. Download the spreadsheet as ``.xlsx`` via File → Download → Microsoft
       Excel (.xlsx). A CSV export only includes the first tab; XLSX is required
       to get the 'AnVIL Prod' sheet.

    3. python3 scripts/prepare_anvil_release.py config <xlsx> <catalog> <first> [<last>]

       Generates the ``<catalog>_sources`` block (e.g. ``anvil15_sources``) from
       the range of rows between the snapshot named ``<first>`` and the snapshot
       named ``<last>`` (both inclusive). If ``<last>`` is omitted, the range
       extends to the last row of the sheet. The script copies its output into
       the clipboard. Paste the output into
       ``deployments/anvilprod/environment.py``. The command is idempotent: it
       always derives ``prev`` as the catalog immediately below ``<catalog>``,
       so running it again after applying the output still produces the same
       block.

    4. python3 scripts/prepare_anvil_release.py column <xlsx> <catalog>

       Generates corrected column values from the ``<catalog>_sources`` block
       added in step 3 and the same spreadsheet. Paste the clipboard content
       into the ``<catalog>`` column. Idempotent: the specific catalog to
       process is chosen by name rather than by picking the last one in
       env.py.
"""
import argparse
import os
import re
import shutil
import subprocess
import sys

import openpyxl
from openpyxl.utils import (
    get_column_letter,
)


def _env_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        '..', 'deployments', 'anvilprod', 'environment.py')


def _copy_to_clipboard(text):
    if sys.platform == 'darwin':
        candidates = [
            ['pbcopy']
        ]
    elif sys.platform.startswith('linux'):
        candidates = [
            ['wl-copy'],
            ['xclip', '-selection', 'clipboard'],
            ['xsel', '--clipboard', '--input']
        ]
    else:
        sys.exit(f'Clipboard copy is not supported on platform {sys.platform!r}')
    for cmd in candidates:
        if shutil.which(cmd[0]) is not None:
            subprocess.run(cmd, input=text.encode(), check=True)
            return
    sys.exit('No clipboard tool found. Install one of: '
             + ', '.join(cmd[0] for cmd in candidates))


def _dataset_name(snapshot):
    if not snapshot.upper().startswith('ANVIL_'):
        snapshot = 'ANVIL_' + snapshot
    return '_'.join(snapshot.split('_')[1:-3])


def _full_snapshot_name(snapshot):
    if not snapshot.upper().startswith('ANVIL_'):
        return 'ANVIL_' + snapshot
    return snapshot


_catalog_re = re.compile(
    r'(anvil\d*?)_sources\s*=\s*union\((\{?\}?|anvil\d*?_sources),'
    r'\s*(\d+),\s*delta\(\[(.*?)\]\)\)',
    re.DOTALL
)
_source_re = re.compile(
    r"source\('([^']+)',\s*'([^']+)'(?:,\s*([\w| ]+))?\)"
)


def _parse_catalogs(env_content, stop_before=None):
    """
    Parse catalog definitions up to (but excluding) ``stop_before``. Returns the
    list of (name, count) tuples and a dict mapping each dataset (lowercased)
    to whether its most recent entry was a pop.
    """
    catalogs = []
    dataset_popped = {}
    for match in _catalog_re.finditer(env_content):
        cat_name = match.group(1)
        if stop_before is not None and cat_name == stop_before:
            break
        count = int(match.group(3))
        catalogs.append((cat_name, count))
        for src in _source_re.finditer(match.group(4)):
            flag_str = src.group(3)
            ds = _dataset_name(src.group(2)).lower()
            dataset_popped[ds] = flag_str is not None and 'pop' in flag_str
    return catalogs, dataset_popped


def cmd_config(args):
    with open(_env_path()) as f:
        env_content = f.read()

    new_name = args.catalog
    m = re.match(r'anvil(\d+)$', new_name)
    if not m:
        sys.exit(f'Invalid catalog name {new_name!r} (expected form: anvilN)')
    new_num = int(m.group(1))
    prev_name = 'anvil' if new_num == 1 else f'anvil{new_num - 1}'

    catalogs, dataset_popped = _parse_catalogs(env_content, stop_before=new_name)
    if not catalogs or catalogs[-1][0] != prev_name:
        sys.exit(f'Previous catalog {prev_name!r} not found in {_env_path()}')
    prev_count = catalogs[-1][1]

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws = wb['AnVIL Prod']

    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    snapshot_col = gp_col = public_col = None
    for i, val in enumerate(header_row):
        if val and 'Snapshot Name' in str(val):
            snapshot_col = i
        if val == 'Google Project ID':
            gp_col = i
        if val == 'Public':
            public_col = i

    if snapshot_col is None:
        sys.exit('Could not find Snapshot Name column')
    if gp_col is None:
        sys.exit('Could not find Google Project ID column')
    if public_col is None:
        sys.exit('Could not find Public column')

    rows = list(ws.iter_rows(min_row=4, values_only=True))
    wb.close()

    first_idx = last_idx = None
    for i, row in enumerate(rows):
        name = row[snapshot_col]
        if name == args.first:
            first_idx = i
        if args.last is not None and name == args.last:
            last_idx = i

    if first_idx is None:
        sys.exit(f'Could not find first snapshot {args.first!r} in sheet')
    if args.last is not None and last_idx is None:
        sys.exit(f'Could not find last snapshot {args.last!r} in sheet')
    if last_idx is None:
        last_idx = len(rows) - 1
    if first_idx > last_idx:
        sys.exit(f'First snapshot {args.first!r} appears after last snapshot {args.last!r}')

    new_entries = []
    for row in rows[first_idx:last_idx + 1]:
        snapshot = row[snapshot_col]
        gp = row[gp_col]
        public = row[public_col]
        if snapshot and gp:
            snapshot = str(snapshot).strip()
            gp = str(gp).strip()
            if snapshot.upper().startswith('ANVIL_'):
                snapshot = snapshot[6:]
            if gp.startswith('datarepo-'):
                gp = gp[9:]
            public = str(public).strip().upper() if public is not None else ''
            if public not in ('Y', 'N'):
                sys.exit(f'Unexpected Public value {public!r} for snapshot {snapshot!r} '
                         f'(expected Y or N)')
            flag = None if public == 'Y' else 'no_ma_mirror'
            new_entries.append((gp, snapshot, flag))

    if not new_entries:
        sys.exit('No valid entries in the specified range')

    delta_entries = []
    new_count = reintroduced_count = updated_count = 0
    for gp, snapshot, flag in new_entries:
        ds = _dataset_name(snapshot)
        ds_lc = ds.lower()
        if ds_lc not in dataset_popped:
            new_count += 1
        elif dataset_popped[ds_lc]:
            reintroduced_count += 1
        else:
            updated_count += 1
        delta_entries.append((gp, snapshot, flag, ds))
    added = new_count + reintroduced_count

    delta_entries.sort(key=lambda e: e[3])

    expected_count = prev_count + added

    source_lines = []
    any_noqa = False
    for gp, snapshot, flag, ds in delta_entries:
        if flag:
            line = f"    source('{gp}', '{snapshot}', {flag}),"
        else:
            line = f"    source('{gp}', '{snapshot}'),"
        if len(line) > 120:
            line += '  # noqa: E501'
            any_noqa = True
        source_lines.append(line)

    result = [f'{new_name}_sources = union({prev_name}_sources, {expected_count}, delta([']
    if any_noqa:
        result.append('    # @formatter:off')
    result.extend(source_lines)
    if any_noqa:
        result.append('    # @formatter:on')
    result.append(']))')

    output = '\n'.join(result)
    _copy_to_clipboard(output)
    print(output, file=sys.stderr)
    print('\n--- Copied to clipboard ---', file=sys.stderr)
    print(f'Previous: {prev_name}_sources (count: {prev_count})', file=sys.stderr)
    print(f'Delta entries: {len(delta_entries)} '
          f'({new_count} new, {reintroduced_count} reintroduced, {updated_count} updated)',
          file=sys.stderr)
    print(f'Expected count: {expected_count}', file=sys.stderr)


def cmd_column(args):
    xlsx_path = args.xlsx
    latest_name = args.catalog

    m = re.match(r'anvil(\d+)$', latest_name)
    if not m:
        sys.exit(f'Invalid catalog name {latest_name!r} (expected form: anvilN)')
    latest_num = int(m.group(1))
    prev_name = 'anvil' if latest_num == 1 else f'anvil{latest_num - 1}'

    with open(_env_path()) as f:
        env_content = f.read()

    latest = None
    for match in _catalog_re.finditer(env_content):
        if match.group(1) == latest_name:
            latest = match
            break
    if latest is None:
        sys.exit(f'Catalog {latest_name!r} not found in {_env_path()}')

    latest_start = latest.start()
    before_latest = env_content[:latest_start]

    all_source_calls = _source_re.findall(before_latest)
    prev_catalog = {}
    for gp, snapshot, flags in all_source_calls:
        ds = _dataset_name(snapshot).lower()
        prev_catalog[ds] = _full_snapshot_name(snapshot)

    latest_delta = latest.group(4)
    delta_sources = _source_re.findall(latest_delta)
    delta_datasets = {_dataset_name(s).lower() for _, s, _ in delta_sources}

    old_snapshots_to_clear = set()
    for ds in delta_datasets:
        if ds in prev_catalog:
            old_snapshots_to_clear.add(prev_catalog[ds])

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['AnVIL Prod']

    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    snapshot_col = None
    latest_col = None
    for i, val in enumerate(header_row):
        if val and 'Snapshot Name' in str(val):
            snapshot_col = i
        if val == latest_name:
            latest_col = i

    if snapshot_col is None:
        sys.exit('Could not find Snapshot Name column')

    if latest_col is None:
        prev_col = None
        for i, val in enumerate(header_row):
            if val == prev_name:
                prev_col = i
        if prev_col is None:
            sys.exit(f'Could not find {latest_name} or {prev_name} column')
        use_col = prev_col
        print(f'Column {latest_name} not found, using {prev_name} as base',
              file=sys.stderr)
    else:
        use_col = latest_col

    output_lines = []
    cleared = 0
    kept = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        snapshot_name = str(row[snapshot_col]) if row[snapshot_col] else None
        current_val = row[use_col]

        if snapshot_name and snapshot_name in old_snapshots_to_clear:
            output_lines.append('')
            if current_val is not None and current_val != '' and current_val != 0:
                cleared += 1
        else:
            if current_val is not None and current_val != '':
                val = current_val
                if isinstance(val, float) and val == int(val):
                    val = int(val)
                output_lines.append(str(val))
                if val == 1:
                    kept += 1
            else:
                output_lines.append('')

    wb.close()

    output = '\n'.join(output_lines)
    _copy_to_clipboard(output)
    paste_cell = f'{get_column_letter(use_col + 1)}4'
    print(f'Catalog: {latest_name} (previous: {prev_name})', file=sys.stderr)
    print(f'Cleared: {cleared}', file=sys.stderr)
    print(f'Kept: {kept}', file=sys.stderr)
    print(f'Total rows: {len(output_lines)}', file=sys.stderr)
    print(f'\n--- Copied to clipboard; select cell {paste_cell} before pasting ---',
          file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    subparsers = parser.add_subparsers(dest='command', required=True)

    config_parser = subparsers.add_parser(
        'config',
        help='Generate anvilXY_sources block from a downloaded spreadsheet.'
    )
    config_parser.add_argument('xlsx', help='Path to the downloaded spreadsheet (.xlsx)')
    config_parser.add_argument('catalog', help='Catalog name to generate (e.g., anvil15)')
    config_parser.add_argument('first', help='Snapshot Name of the first row to include')
    config_parser.add_argument('last', nargs='?', default=None,
                               help='Snapshot Name of the last row to include (default: end of sheet)')
    config_parser.set_defaults(func=cmd_config)

    column_parser = subparsers.add_parser(
        'column',
        help='Generate corrected anvilXY column values from a downloaded spreadsheet.'
    )
    column_parser.add_argument('xlsx', help='Path to the downloaded spreadsheet (.xlsx)')
    column_parser.add_argument('catalog', help='Catalog name to process (e.g., anvil15)')
    column_parser.set_defaults(func=cmd_column)

    args = parser.parse_args()
    args.func(args)


if __name__ == '__main__':
    main()
