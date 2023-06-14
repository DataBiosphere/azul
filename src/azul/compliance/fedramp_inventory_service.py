from abc import (
    ABCMeta,
    abstractmethod,
)
from collections import (
    Counter,
    defaultdict,
)
import inspect
import json
import logging
from operator import (
    attrgetter,
)
import pathlib
import sys
from typing import (
    AbstractSet,
    Iterable,
    Iterator,
    Optional,
    Sequence,
)

import attr
import attrs
from furl import (
    furl,
)
import gitlab.v4.objects.projects
from more_itertools import (
    chunked,
    flatten,
)
import openpyxl
from openpyxl.utils import (
    get_column_letter,
)
from openpyxl.worksheet.worksheet import (
    Worksheet,
)

from azul import (
    cached_property,
    config,
)
from azul.deployment import (
    aws,
)
from azul.types import (
    JSON,
    JSONs,
)

log = logging.getLogger(__name__)


@attr.s(auto_attribs=True, frozen=True, kw_only=True)
class ResourceConfig:
    id: str
    name: Optional[str]
    region: str
    type: str
    state_id: str
    config: JSON
    supplementary_config: JSON

    @classmethod
    def from_response(cls, response: dict) -> 'ResourceConfig':
        return cls(
            id=response['resourceId'],
            name=response.get('resourceName'),
            region=response['awsRegion'],
            type=response['resourceType'],
            state_id=response['configurationStateId'],
            config=json.loads(response['configuration']),
            supplementary_config=response['supplementaryConfiguration']
        )


null_str = Optional[str]


class YesNo:
    yes = 'Yes'
    no = 'No'

    @classmethod
    def from_bool(cls, b: bool) -> str:
        return cls.yes if b else cls.no


@attr.s(auto_attribs=True, frozen=True, kw_only=True)
class InventoryRow:
    unique_id: null_str = attr.ib(default=None)
    ip_address: null_str = attr.ib(default=None)
    is_virtual: null_str = attr.ib(default=None)
    is_public: null_str = attr.ib(default=None)
    dns_name: null_str = attr.ib(default=None)
    netbios_name: null_str = attr.ib(default=None)
    mac_address: null_str = attr.ib(default=None)
    authenticated_scan_planned: null_str = attr.ib(default=None)
    baseline_config: null_str = attr.ib(default=None)
    os: null_str = attr.ib(default=None)
    location: null_str = attr.ib(default=None)
    asset_type: null_str = attr.ib(default=None)
    hardware_model: null_str = attr.ib(default=None)
    in_latest_scan: null_str = attr.ib(default=None)
    software_vendor: null_str = attr.ib(default=None)
    software_product_name: null_str = attr.ib(default=None)
    patch_level: null_str = attr.ib(default=None)
    purpose: null_str = attr.ib(default=None)
    comments: null_str = attr.ib(default=None)
    asset_tag: null_str = attr.ib(default=None)
    network_id: null_str = attr.ib(default=None)
    system_owner: null_str = attr.ib(default=None)
    application_owner: null_str = attr.ib(default=None)


class Mapper(metaclass=ABCMeta):

    @abstractmethod
    def map(self, resource: ResourceConfig) -> Iterable[InventoryRow]:
        raise NotImplementedError

    def _common_fields(self, resource: ResourceConfig, *, id_suffix: Optional[str] = None) -> dict:
        return dict(
            asset_tag=resource.name,
            location=resource.region,
            software_vendor='AWS',
            system_owner=config.owner,
            unique_id=resource.id + ('' if id_suffix is None else f'/{id_suffix}')
        )

    def _supported_resource_types(self) -> AbstractSet[str]:
        return frozenset()

    def can_map(self, resource: ResourceConfig) -> bool:
        return resource.type in self._supported_resource_types()


class LambdaMapper(Mapper):

    def _supported_resource_types(self) -> set[str]:
        return {'AWS::Lambda::Function'}

    def map(self, resource: ResourceConfig) -> Iterator[InventoryRow]:
        yield InventoryRow(
            asset_type='AWS Lambda Function',
            baseline_config=resource.config['runtime'],
            is_public=YesNo.no,
            is_virtual=YesNo.yes,
            purpose=resource.config.get('description'),
            software_product_name='AWS Lambda',
            **self._common_fields(resource)
        )


class ElasticSearchMapper(Mapper):

    def _supported_resource_types(self) -> set[str]:
        return {'AWS::Elasticsearch::Domain'}

    def map(self, resource: ResourceConfig) -> Iterator[InventoryRow]:
        yield InventoryRow(
            asset_type='AWS OpenSearch Domain',
            baseline_config=resource.config['elasticsearchVersion'],
            is_public=YesNo.no,
            is_virtual=YesNo.yes,
            network_id=resource.config['endpoints'].get('vpc'),
            patch_level=resource.config.get('serviceSoftwareOptions', {}).get('currentVersion'),
            software_product_name='AWS OpenSearch',
            **self._common_fields(resource)
        )


class EC2Mapper(Mapper):

    def _supported_resource_types(self) -> set[str]:
        return {'AWS::EC2::Instance'}

    def map(self, resource: ResourceConfig) -> Iterable[InventoryRow]:
        for nic in resource.config['networkInterfaces']:
            for ip_addresses in nic['privateIpAddresses']:
                ip_addresses: JSON
                association = ip_addresses.get('association')
                ips = [
                    dict(ip_address=ip_addresses['privateIpAddress'],
                         dns_name=ip_addresses['privateDnsName'],
                         is_public=YesNo.no),
                    *(() if association is None else (
                        dict(ip_address=association['publicIp'],
                             is_public=YesNo.yes,
                             dns_name=resource.config.get('publicDnsName'))
                    ))
                ]
                for ip_fields in ips:
                    yield InventoryRow(
                        asset_type='AWS EC2 Instance',
                        authenticated_scan_planned=YesNo.yes,
                        baseline_config=resource.config['imageId'],
                        hardware_model=resource.config['instanceType'],
                        is_virtual=YesNo.yes,
                        mac_address=nic['macAddress'],
                        network_id=ip_addresses.get('subnetId'),
                        **ip_fields,
                        **self._common_fields(resource, id_suffix=ip_fields['ip_address'])
                    )

    def _get_ip_address(self, ip_addresses: JSON, keys) -> str:
        for key in keys:
            ip_addresses = ip_addresses[key]
        return ip_addresses


class ELBMapper(Mapper):

    def _supported_resource_types(self) -> set[str]:
        return {
            'AWS::ElasticLoadBalancing::LoadBalancer',
            'AWS::ElasticLoadBalancingV2::LoadBalancer'
        }

    def map(self, resource: ResourceConfig) -> Iterator[InventoryRow]:
        ip_addresses = self._get_ip_addresses(resource.config['availabilityZones'])
        if not ip_addresses:
            ip_addresses = [None]
        for ip_address in ip_addresses:
            yield InventoryRow(
                dns_name=resource.config['dNSName'],
                ip_address=ip_address,
                is_public=YesNo.from_bool(resource.config['scheme'] == 'internet-facing'),
                is_virtual=YesNo.yes,
                **self._polymorphic_fields(resource),
                **self._common_fields(resource, id_suffix=ip_address)
            )

    def _polymorphic_fields(self, resource: ResourceConfig) -> dict[str, str]:
        # Classic ELBs have key of 'vpcid' while V2 ELBs have key of 'vpcId'
        prefix = 'AWS Elastic Load Balancer-'
        if resource.type == 'AWS::ElasticLoadBalancing::LoadBalancer':
            asset_type = prefix + 'Classic'
            network_id = resource.config['vpcid']
        else:
            asset_type = prefix + resource.config['type']
            network_id = resource.config['vpcId']
        return dict(asset_type=asset_type, network_id=network_id)

    def _get_ip_addresses(self, availability_zones: JSONs) -> set[Optional[str]]:
        return {
            load_balancer_address.get('ipAddress')
            for availability_zone in availability_zones
            for load_balancer_addresses in availability_zone.get('loadBalancerAddresses', ())
            for load_balancer_address in load_balancer_addresses
        }


class NetworkInterfaceMapper(Mapper):

    def _supported_resource_types(self) -> AbstractSet[str]:
        return {'AWS::EC2::NetworkInterface'}

    def map(self, resource: ResourceConfig) -> Iterable[InventoryRow]:
        ips = [
            dict(is_public=YesNo.no,
                 ip_address=private_ip['privateIpAddress'],
                 dns_name=private_ip.get('privateDnsName'))
            for private_ip in resource.config['privateIpAddresses']
        ]
        association = resource.config.get('association')
        if association is not None:
            ips.append(dict(is_public=YesNo.yes,
                            ip_address=association['publicIp'],
                            dns_name=association['publicDnsName']))
        for ip_fields in ips:
            yield InventoryRow(
                asset_type='AWS EC2 Network Interface',
                mac_address=resource.config.get('macAddress'),
                network_id=resource.config['subnetId'],
                purpose=resource.config.get('description'),
                **ip_fields,
                **self._common_fields(resource, id_suffix=ip_fields['ip_address'])
            )


class S3Mapper(Mapper):

    def _supported_resource_types(self) -> set[str]:
        return {'AWS::S3::Bucket'}

    def map(self, resource: ResourceConfig) -> Iterator[InventoryRow]:
        yield InventoryRow(
            asset_type='AWS S3 Bucket',
            comments=self._get_encryption_status(resource),
            is_public=YesNo.from_bool(self._get_is_public(resource)),
            is_virtual=YesNo.yes,
            **self._common_fields(resource)
        )

    def _get_is_public(self, resource: ResourceConfig) -> bool:
        try:
            public_access_config = resource.supplementary_config['PublicAccessBlockConfiguration']
        except KeyError:
            # If there is no PublicAccessBlockConfiguration then this bucket is public
            return True
        else:
            public_access_config = json.loads(public_access_config)
            # The bucket is public if any access blocks are false
            return not all(public_access_config.values())

    def _get_encryption_status(self, resource: ResourceConfig) -> str:
        if 'ServerSideEncryptionConfiguration' in resource.supplementary_config:
            return 'Encrypted'
        else:
            return 'Not encrypted'


class DynamoDbTableMapper(Mapper):

    def _supported_resource_types(self) -> set[str]:
        return {'AWS::DynamoDB::Table'}

    def map(self, resource: ResourceConfig) -> Iterator[InventoryRow]:
        yield InventoryRow(
            asset_type='AWS DynamoDB Table',
            is_public=YesNo.no,
            is_virtual=YesNo.yes,
            software_product_name='AWS DynamoDB',
            **self._common_fields(resource)
        )


class ElasticIPMapper(Mapper):

    def _supported_resource_types(self) -> AbstractSet[str]:
        return {'AWS::EC2::EIP'}

    def map(self, resource: ResourceConfig) -> Iterable[InventoryRow]:
        for ip, is_public in [
            (resource.config['publicIp'], YesNo.yes),
            (resource.config['privateIpAddress'], YesNo.no)
        ]:
            yield InventoryRow(
                asset_type='AWS EC2 Elastic IP',
                ip_address=ip,
                is_public=is_public,
                network_id=resource.config['networkInterfaceId'],
                **self._common_fields(resource, id_suffix=ip)
            )


class RDSMapper(Mapper):

    def _supported_resource_types(self) -> set[str]:
        return {'AWS::RDS::DBInstance'}

    def map(self, resource: ResourceConfig) -> Iterator[InventoryRow]:
        yield InventoryRow(
            asset_type='AWS RDS Instance',
            hardware_model=resource.config['dBInstanceClass'],
            is_public=YesNo.from_bool(resource.config['publiclyAccessible']),
            is_virtual=YesNo.yes,
            network_id=resource.config.get('dBSubnetGroup', {}).get('vpcId'),
            software_product_name=f"{resource.config['engine']}-{resource.config['engineVersion']}",
            **self._common_fields(resource)
        )


class VPCMapper(Mapper):

    def _supported_resource_types(self) -> set[str]:
        return {'AWS::EC2::VPC'}

    def map(self, resource: ResourceConfig) -> Iterator[InventoryRow]:
        yield InventoryRow(
            asset_type='AWS VPC',
            baseline_config=resource.state_id,
            ip_address=resource.config['cidrBlock'],
            is_public=YesNo.yes,
            is_virtual=YesNo.yes,
            network_id=resource.config['vpcId'],
            **self._common_fields(resource)
        )


class ACMCertificateMapper(Mapper):

    def _supported_resource_types(self) -> AbstractSet[str]:
        return {'AWS::ACM::Certificate'}

    def map(self, resource: ResourceConfig) -> Iterable[InventoryRow]:
        yield InventoryRow(
            asset_type='AWS ACM Certificate',
            **self._common_fields(resource)
        )
        for user in resource.config['inUseBy']:
            parts, id = user.split('/', 1)
            parts = parts.split(':')
            if parts[:2] == ['aws', 'clientvpn']:
                _, resource_type, region, stage = parts
                url = '.'.join([id, stage, resource_type, region, 'amazonaws.com'])
                yield InventoryRow(
                    asset_tag=user,
                    asset_type='AWS Client VPN',
                    dns_name=url,
                    location=region,
                    software_vendor='AWS',
                    unique_id=url + ':443',
                )


class ResourceComplianceMapper(Mapper):

    def _supported_resource_types(self) -> AbstractSet[str]:
        return {'AWS::Config::ResourceCompliance'}

    def map(self, resource: ResourceConfig) -> Iterable[InventoryRow]:
        # Intentionally omit rows for this resource type
        return ()


class DefaultMapper(Mapper):

    def can_map(self, resource: ResourceConfig) -> bool:
        return True

    def map(self, resource: ResourceConfig) -> Iterable[InventoryRow]:
        yield InventoryRow(
            asset_type=resource.type,
            **self._common_fields(resource)
        )


class FedRAMPInventoryService:
    default_column_width = 10
    first_writable_row = 6
    report_worksheet_name = 'Inventory'

    @property
    def config(self):
        return aws.client('config')

    @cached_property
    def _mappers(self) -> Sequence[Mapper]:
        current_module = sys.modules[__name__]

        def is_mapper_cls(o: object) -> bool:
            return (
                inspect.isclass(o)
                and not inspect.isabstract(o)
                and issubclass(o, Mapper)
            )

        mapper_clss = [
            mapper_cls
            for name, mapper_cls in inspect.getmembers(current_module, is_mapper_cls)
        ]

        def get_linenno(o: type) -> int:
            src, lineno = inspect.findsource(o)
            return lineno

        mapper_clss.sort(key=get_linenno)
        return [mapper_cls() for mapper_cls in mapper_clss]

    def resource_ids_by_type(self) -> defaultdict[str, list[str]]:
        resource_ids_by_type = defaultdict(list)
        for resource_type in self._all_aws_resource_types:
            args = dict(resourceType=resource_type)
            while True:
                response = self.config.list_discovered_resources(**args)
                resources = response['resourceIdentifiers']
                log.debug('Discovered %d resources of type %s', len(resources), resource_type)
                for resource in resources:
                    assert resource['resourceType'] == resource_type
                    resource_ids_by_type[resource_type].append(resource['resourceId'])
                next_token = response.get('nextToken')
                if next_token is None:
                    break
                else:
                    args['nextToken'] = next_token
        return resource_ids_by_type

    def get_resources(self) -> Iterator[ResourceConfig]:
        for resource_type, resource_ids in self.resource_ids_by_type().items():
            # Maximum permitted batch size
            for resource_ids in chunked(resource_ids, 100):
                resource_keys = [
                    dict(resourceType=resource_type, resourceId=resource_id)
                    for resource_id in resource_ids
                ]
                while resource_keys:
                    response = self.config.batch_get_resource_config(resourceKeys=resource_keys)
                    items = response['baseConfigurationItems']
                    log.debug('Got page of %d resources of type %s', len(items), resource_type)
                    yield from map(ResourceConfig.from_response, items)
                    resource_keys = response['unprocessedResourceKeys']

    def get_inventory(self, resources: Iterable[ResourceConfig]) -> Iterable[InventoryRow]:
        rows_by_mapper: defaultdict[Mapper, list[InventoryRow]] = defaultdict(list)
        resource_counts = Counter()
        row_counts = Counter()
        for resource in resources:
            mapper = self._get_mapper(resource)
            log.debug('Mapping %r resource using %r',
                      resource.type, type(mapper).__name__)
            rows = sorted(mapper.map(resource), key=attrgetter('unique_id'))
            log.debug('Mapped to %d rows', len(rows))
            resource_counts[resource.type] += 1
            row_counts[resource.type] += len(rows)
            rows_by_mapper[mapper].extend(rows)

        log.info('Inventory contents:')
        print(f'\n{"Resource type":<42s}'
              f'{"# resources":<20s}'
              f'{"# rows":<20s}\n')
        for resource_type in resource_counts.keys():
            print(f'{resource_type:<42s}'
                  f'{resource_counts[resource_type]:>15d}'
                  f'{row_counts[resource_type]:>10d}')

        return flatten(rows_by_mapper[mapper] for mapper in self._mappers)

    def get_synthetic_inventory(self) -> Iterable[InventoryRow]:
        data_browser_url = furl(scheme='https', netloc=config.data_browser_domain)
        yield InventoryRow(
            asset_type='Application endpoint',
            dns_name=str(data_browser_url),
            is_public=YesNo.yes,
            purpose='UI for external users',
            software_vendor='UCSC',
            system_owner=config.owner,
            unique_id='Data Browser UI',
        )
        yield InventoryRow(
            asset_type='Service endpoint',
            dns_name=str(config.service_endpoint),
            is_public=YesNo.from_bool(not config.private_api),
            purpose='Service API (backend for Data Browser UI, programmatic use by external users)',
            software_vendor='UCSC',
            system_owner=config.owner,
            unique_id='Service REST API',
        )
        yield InventoryRow(
            asset_type='Application endpoint',
            dns_name=str(config.indexer_endpoint),
            is_public=YesNo.from_bool(not config.private_api),
            purpose='Indexer API (primarily for internal users)',
            software_vendor='UCSC',
            system_owner=config.owner,
            unique_id='Indexer API',
        )

        for unique_id, purpose, port, scheme in [
            ('GitLab UI', 'CI/CD (internal users only)', None, 'https'),
            ('GitLab SSH', 'CI/CD (system administrators only)', 2222, 'ssh'),
            ('GitLab Git', 'Source repository for CI/CD (internal users only)', 22, 'git+ssh')
        ]:
            gitlab_url = furl(scheme=scheme,
                              host=f'gitlab.{config.domain_name}',
                              port=port)
            yield InventoryRow(
                asset_type='Service endpoint',
                dns_name=str(gitlab_url),
                is_public=YesNo.no,
                software_vendor='GitLab',
                system_owner=config.owner,
                purpose=purpose,
                unique_id=unique_id,
            )

    def write_report(self,
                     inventory: Iterable[InventoryRow],
                     template_path: pathlib.Path,
                     output_path: pathlib.Path
                     ) -> None:
        workbook = openpyxl.load_workbook(template_path)
        worksheet = workbook[self.report_worksheet_name]
        for row_number, row in enumerate(inventory, start=self.first_writable_row):
            row = attr.astuple(row)
            for column_number, value in enumerate(row, start=1):
                self._write_cell_if_value_provided(worksheet,
                                                   column=column_number,
                                                   row=row_number,
                                                   value=value)
        workbook.save(output_path)
        log.info('Wrote report to %s', output_path)

    def update_wiki(self,
                    project: gitlab.v4.objects.projects.Project,
                    page_name: str,
                    resources: Iterable[ResourceConfig],
                    ) -> None:
        content = self._wiki_content(resources)
        try:
            page = project.wikis.get(page_name)
        except gitlab.exceptions.GitlabError as e:
            if e.response_code == 404:
                log.info('Wiki page %r not found', page_name)
                project.wikis.create({
                    'title': page_name,
                    'content': content
                })
                log.info('Created wiki page %r (character count: %d)',
                         page_name, len(content))
            else:
                raise
        else:
            old_length = len(page.content)
            page.content = content
            page.save()
            log.info('Updated wiki page %r (character count: %d -> %d)',
                     page_name, old_length, len(content))

    def _get_mapper(self, resource: ResourceConfig) -> Mapper:
        return next(
            mapper
            for mapper in self._mappers
            if mapper.can_map(resource)
        )

    def _write_cell_if_value_provided(self,
                                      worksheet: Worksheet,
                                      column: int,
                                      row: int,
                                      value: Optional[str]
                                      ) -> None:
        if value:
            # Scale the size of the column with the input value if necessary.
            # By default, width is None.
            dimensions = worksheet.column_dimensions[get_column_letter(column)]
            if dimensions.width is None:
                dimensions.width = self.default_column_width
            else:
                dimensions.width = max(dimensions.width, len(value))
            worksheet.cell(column=column, row=row, value=value)

    def _wiki_content(self, resources: Iterable[ResourceConfig]) -> str:
        return '\n\n'.join(
            f'```\n{json.dumps(attrs.asdict(resource), indent=4)}\n```'
            for resource in resources
        )

    # https://docs.aws.amazon.com/config/latest/APIReference/API_ListDiscoveredResources.html#API_ListDiscoveredResources_RequestSyntax
    _all_aws_resource_types = [
        'AWS::ACM::Certificate',
        'AWS::AccessAnalyzer::Analyzer',
        'AWS::AmazonMQ::Broker',
        'AWS::ApiGateway::RestApi',
        'AWS::ApiGateway::Stage',
        'AWS::ApiGatewayV2::Api',
        'AWS::ApiGatewayV2::Stage',
        'AWS::AppConfig::Application',
        'AWS::AppConfig::ConfigurationProfile',
        'AWS::AppConfig::Environment',
        'AWS::AppStream::DirectoryConfig',
        'AWS::AppSync::GraphQLApi',
        'AWS::Athena::DataCatalog',
        'AWS::Athena::WorkGroup',
        'AWS::AutoScaling::AutoScalingGroup',
        'AWS::AutoScaling::LaunchConfiguration',
        'AWS::AutoScaling::ScalingPolicy',
        'AWS::AutoScaling::ScheduledAction',
        'AWS::AutoScaling::WarmPool',
        'AWS::Backup::BackupPlan',
        'AWS::Backup::BackupSelection',
        'AWS::Backup::BackupVault',
        'AWS::Backup::RecoveryPoint',
        'AWS::Backup::ReportPlan',
        'AWS::Batch::ComputeEnvironment',
        'AWS::Batch::JobQueue',
        'AWS::Budgets::BudgetsAction',
        'AWS::Cloud9::EnvironmentEC2',
        'AWS::CloudFormation::Stack',
        'AWS::CloudFront::Distribution',
        'AWS::CloudFront::StreamingDistribution',
        'AWS::CloudTrail::Trail',
        'AWS::CloudWatch::Alarm',
        'AWS::CodeBuild::Project',
        'AWS::CodeDeploy::Application',
        'AWS::CodeDeploy::DeploymentConfig',
        'AWS::CodeDeploy::DeploymentGroup',
        'AWS::CodeGuruReviewer::RepositoryAssociation',
        'AWS::CodePipeline::Pipeline',
        'AWS::Config::ConformancePackCompliance',
        'AWS::Config::ResourceCompliance',
        'AWS::Connect::PhoneNumber',
        'AWS::CustomerProfiles::Domain',
        'AWS::DMS::Certificate',
        'AWS::DMS::EventSubscription',
        'AWS::DMS::ReplicationSubnetGroup',
        'AWS::DataSync::LocationEFS',
        'AWS::DataSync::LocationFSxLustre',
        'AWS::DataSync::LocationFSxWindows',
        'AWS::DataSync::LocationHDFS',
        'AWS::DataSync::LocationNFS',
        'AWS::DataSync::LocationObjectStorage',
        'AWS::DataSync::LocationS3',
        'AWS::DataSync::LocationSMB',
        'AWS::DataSync::Task',
        'AWS::Detective::Graph',
        'AWS::DeviceFarm::TestGridProject',
        'AWS::DynamoDB::Table',
        'AWS::EC2::CustomerGateway',
        'AWS::EC2::DHCPOptions',
        'AWS::EC2::EIP',
        'AWS::EC2::EgressOnlyInternetGateway',
        'AWS::EC2::FlowLog',
        'AWS::EC2::Host',
        'AWS::EC2::IPAM',
        'AWS::EC2::Instance',
        'AWS::EC2::InternetGateway',
        'AWS::EC2::LaunchTemplate',
        'AWS::EC2::NatGateway',
        'AWS::EC2::NetworkAcl',
        'AWS::EC2::NetworkInsightsAccessScopeAnalysis',
        'AWS::EC2::NetworkInsightsPath',
        'AWS::EC2::NetworkInterface',
        'AWS::EC2::RegisteredHAInstance',
        'AWS::EC2::RouteTable',
        'AWS::EC2::SecurityGroup',
        'AWS::EC2::Subnet',
        'AWS::EC2::TrafficMirrorFilter',
        'AWS::EC2::TrafficMirrorSession',
        'AWS::EC2::TrafficMirrorTarget',
        'AWS::EC2::TransitGateway',
        'AWS::EC2::TransitGatewayAttachment',
        'AWS::EC2::TransitGatewayRouteTable',
        'AWS::EC2::VPC',
        'AWS::EC2::VPCEndpoint',
        'AWS::EC2::VPCEndpointService',
        'AWS::EC2::VPCPeeringConnection',
        'AWS::EC2::VPNConnection',
        'AWS::EC2::VPNGateway',
        'AWS::EC2::Volume',
        'AWS::ECR::PublicRepository',
        'AWS::ECR::RegistryPolicy',
        'AWS::ECR::Repository',
        'AWS::ECS::Cluster',
        'AWS::ECS::Service',
        'AWS::ECS::TaskDefinition',
        'AWS::EFS::AccessPoint',
        'AWS::EFS::FileSystem',
        'AWS::EKS::Addon',
        'AWS::EKS::Cluster',
        'AWS::EKS::FargateProfile',
        'AWS::EKS::IdentityProviderConfig',
        'AWS::EMR::SecurityConfiguration',
        'AWS::ElasticBeanstalk::Application',
        'AWS::ElasticBeanstalk::ApplicationVersion',
        'AWS::ElasticBeanstalk::Environment',
        'AWS::ElasticLoadBalancing::LoadBalancer',
        'AWS::ElasticLoadBalancingV2::Listener',
        'AWS::ElasticLoadBalancingV2::LoadBalancer',
        'AWS::Elasticsearch::Domain',
        'AWS::EventSchemas::Discoverer',
        'AWS::EventSchemas::Registry',
        'AWS::EventSchemas::RegistryPolicy',
        'AWS::EventSchemas::Schema',
        'AWS::Events::ApiDestination',
        'AWS::Events::Archive',
        'AWS::Events::Connection',
        'AWS::Events::Endpoint',
        'AWS::Events::EventBus',
        'AWS::Events::Rule',
        'AWS::FIS::ExperimentTemplate',
        'AWS::FraudDetector::EntityType',
        'AWS::FraudDetector::Label',
        'AWS::FraudDetector::Outcome',
        'AWS::FraudDetector::Variable',
        'AWS::GlobalAccelerator::Accelerator',
        'AWS::GlobalAccelerator::EndpointGroup',
        'AWS::GlobalAccelerator::Listener',
        'AWS::Glue::Classifier',
        'AWS::Glue::Job',
        'AWS::Glue::MLTransform',
        'AWS::GuardDuty::Detector',
        'AWS::GuardDuty::Filter',
        'AWS::GuardDuty::IPSet',
        'AWS::GuardDuty::ThreatIntelSet',
        'AWS::HealthLake::FHIRDatastore',
        'AWS::IAM::Group',
        'AWS::IAM::Policy',
        'AWS::IAM::Role',
        'AWS::IAM::User',
        'AWS::IVS::Channel',
        'AWS::IVS::PlaybackKeyPair',
        'AWS::IVS::RecordingConfiguration',
        'AWS::ImageBuilder::ContainerRecipe',
        'AWS::ImageBuilder::DistributionConfiguration',
        'AWS::ImageBuilder::InfrastructureConfiguration',
        'AWS::IoT::AccountAuditConfiguration',
        'AWS::IoT::Authorizer',
        'AWS::IoT::CustomMetric',
        'AWS::IoT::Dimension',
        'AWS::IoT::MitigationAction',
        'AWS::IoT::Policy',
        'AWS::IoT::RoleAlias',
        'AWS::IoT::ScheduledAudit',
        'AWS::IoT::SecurityProfile',
        'AWS::IoTAnalytics::Channel',
        'AWS::IoTAnalytics::Dataset',
        'AWS::IoTAnalytics::Datastore',
        'AWS::IoTAnalytics::Pipeline',
        'AWS::IoTEvents::AlarmModel',
        'AWS::IoTEvents::DetectorModel',
        'AWS::IoTEvents::Input',
        'AWS::IoTSiteWise::AssetModel',
        'AWS::IoTSiteWise::Dashboard',
        'AWS::IoTSiteWise::Gateway',
        'AWS::IoTSiteWise::Portal',
        'AWS::IoTSiteWise::Project',
        'AWS::IoTTwinMaker::Entity',
        'AWS::IoTTwinMaker::Scene',
        'AWS::IoTTwinMaker::Workspace',
        'AWS::KMS::Key',
        'AWS::Kinesis::Stream',
        'AWS::Kinesis::StreamConsumer',
        'AWS::KinesisAnalyticsV2::Application',
        'AWS::KinesisVideo::SignalingChannel',
        'AWS::Lambda::Function',
        'AWS::Lex::Bot',
        'AWS::Lex::BotAlias',
        'AWS::Lightsail::Bucket',
        'AWS::Lightsail::Certificate',
        'AWS::Lightsail::Disk',
        'AWS::Lightsail::StaticIp',
        'AWS::LookoutMetrics::Alert',
        'AWS::LookoutVision::Project',
        'AWS::MSK::Cluster',
        'AWS::MediaPackage::PackagingConfiguration',
        'AWS::MediaPackage::PackagingGroup',
        'AWS::NetworkFirewall::Firewall',
        'AWS::NetworkFirewall::FirewallPolicy',
        'AWS::NetworkFirewall::RuleGroup',
        'AWS::NetworkManager::TransitGatewayRegistration',
        'AWS::OpenSearch::Domain',
        'AWS::Pinpoint::ApplicationSettings',
        'AWS::Pinpoint::Segment',
        'AWS::QLDB::Ledger',
        'AWS::RDS::DBCluster',
        'AWS::RDS::DBClusterSnapshot',
        'AWS::RDS::DBInstance',
        'AWS::RDS::DBSecurityGroup',
        'AWS::RDS::DBSnapshot',
        'AWS::RDS::DBSubnetGroup',
        'AWS::RDS::EventSubscription',
        'AWS::RDS::GlobalCluster',
        'AWS::RUM::AppMonitor',
        'AWS::Redshift::Cluster',
        'AWS::Redshift::ClusterParameterGroup',
        'AWS::Redshift::ClusterSecurityGroup',
        'AWS::Redshift::ClusterSnapshot',
        'AWS::Redshift::ClusterSubnetGroup',
        'AWS::Redshift::EventSubscription',
        'AWS::ResilienceHub::ResiliencyPolicy',
        'AWS::RoboMaker::RobotApplication',
        'AWS::RoboMaker::RobotApplicationVersion',
        'AWS::RoboMaker::SimulationApplication',
        'AWS::Route53::HostedZone',
        'AWS::Route53RecoveryControl::Cluster',
        'AWS::Route53RecoveryControl::ControlPanel',
        'AWS::Route53RecoveryControl::RoutingControl',
        'AWS::Route53RecoveryControl::SafetyRule',
        'AWS::Route53RecoveryReadiness::Cell',
        'AWS::Route53RecoveryReadiness::ReadinessCheck',
        'AWS::Route53RecoveryReadiness::RecoveryGroup',
        'AWS::Route53RecoveryReadiness::ResourceSet',
        'AWS::Route53Resolver::FirewallDomainList',
        'AWS::Route53Resolver::ResolverEndpoint',
        'AWS::Route53Resolver::ResolverRule',
        'AWS::Route53Resolver::ResolverRuleAssociation',
        'AWS::S3::AccountPublicAccessBlock',
        'AWS::S3::Bucket',
        'AWS::S3::MultiRegionAccessPoint',
        'AWS::S3::StorageLens',
        'AWS::SES::ConfigurationSet',
        'AWS::SES::ContactList',
        'AWS::SES::ReceiptFilter',
        'AWS::SES::ReceiptRuleSet',
        'AWS::SES::Template',
        'AWS::SNS::Topic',
        'AWS::SQS::Queue',
        'AWS::SSM::AssociationCompliance',
        'AWS::SSM::FileData',
        'AWS::SSM::ManagedInstanceInventory',
        'AWS::SSM::PatchCompliance',
        'AWS::SageMaker::CodeRepository',
        'AWS::SageMaker::Model',
        'AWS::SageMaker::NotebookInstanceLifecycleConfig',
        'AWS::SageMaker::Workteam',
        'AWS::SecretsManager::Secret',
        'AWS::ServiceCatalog::CloudFormationProduct',
        'AWS::ServiceCatalog::CloudFormationProvisionedProduct',
        'AWS::ServiceCatalog::Portfolio',
        'AWS::ServiceDiscovery::HttpNamespace',
        'AWS::ServiceDiscovery::PublicDnsNamespace',
        'AWS::ServiceDiscovery::Service',
        'AWS::Shield::Protection',
        'AWS::ShieldRegional::Protection',
        'AWS::StepFunctions::Activity',
        'AWS::StepFunctions::StateMachine',
        'AWS::Transfer::Workflow',
        'AWS::WAF::RateBasedRule',
        'AWS::WAF::Rule',
        'AWS::WAF::RuleGroup',
        'AWS::WAF::WebACL',
        'AWS::WAFRegional::RateBasedRule',
        'AWS::WAFRegional::Rule',
        'AWS::WAFRegional::RuleGroup',
        'AWS::WAFRegional::WebACL',
        'AWS::WAFv2::IPSet',
        'AWS::WAFv2::ManagedRuleSet',
        'AWS::WAFv2::RegexPatternSet',
        'AWS::WAFv2::RuleGroup',
        'AWS::WAFv2::WebACL',
        'AWS::WorkSpaces::ConnectionAlias',
        'AWS::WorkSpaces::Workspace',
        'AWS::XRay::EncryptionConfig',
    ]
